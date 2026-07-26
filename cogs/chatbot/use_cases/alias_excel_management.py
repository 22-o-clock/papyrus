import datetime
import io
import uuid
from typing import Any
from zipfile import BadZipFile

import discord
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from cogs.chatbot.constants import (
    MEMBER_ALIAS_ACTION_LABELS,
    MEMBER_ALIAS_EVIDENCE_COLUMN,
    MEMBER_ALIAS_HEADERS,
    MEMBER_ALIAS_MEMBER_SHEET_NAME,
    MEMBER_ALIAS_SHEET_NAME,
    MEMBER_ALIAS_STATUS_LABELS,
)
from cogs.chatbot.repositories.member_alias import (
    ChatbotMemberAliasRepository,
    MemberAliasReviewRecord,
    MemberAliasUpdateInput,
)


class AliasExcelManagementUseCases:
    """Chatbotのメンバー別名をレビューするExcelを入出力します。"""

    def __init__(self, session_factory: async_sessionmaker[AsyncSession]) -> None:
        self.member_alias_repository = ChatbotMemberAliasRepository(session_factory)

    async def export_chatbot_member_aliases(self, interaction: discord.Interaction) -> None:
        """管理者向けに全別名と根拠を一括編集用Excelで返します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("別名の出力には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        if interaction.guild is None:
            await interaction.response.send_message("サーバー内で実行してください。", ephemeral=True)
            return
        records = await self.member_alias_repository.get_review_records()
        if not records:
            await interaction.response.send_message("保存されているメンバー別名はありません。", ephemeral=True)
            return
        member_labels = {
            member.id: f"{member.display_name} ({member.id})"
            for member in sorted(interaction.guild.members, key=lambda item: (item.display_name.casefold(), item.id))
        }
        workbook = self._build_workbook(records, member_labels, interaction.guild.id)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
        await interaction.response.send_message(
            f"メンバー別名 {len(records)} 件を出力しました。",
            ephemeral=True,
            file=discord.File(output, filename=f"chatbot_member_aliases_{timestamp}.xlsx"),
        )

    async def import_chatbot_member_aliases(
        self,
        interaction: discord.Interaction,
        attachment: discord.Attachment,
    ) -> None:
        """全行を検証し、問題がなければ別名変更を一括適用します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("別名の取込には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        if interaction.guild is None:
            await interaction.response.send_message("サーバー内で実行してください。", ephemeral=True)
            return
        try:
            workbook = load_workbook(io.BytesIO(await attachment.read()), data_only=True)
            updates = self._parse_workbook(workbook, interaction.guild)
            await self.member_alias_repository.apply_updates(updates, interaction.user.id)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            SQLAlchemyError,
            TypeError,
            ValueError,
        ) as error:
            await interaction.response.send_message(f"取り込みを中止しました: {error}", ephemeral=True)
            return
        await interaction.response.send_message(f"メンバー別名 {len(updates)} 件を一括反映しました。", ephemeral=True)

    def _build_workbook(
        self,
        records: list[MemberAliasReviewRecord],
        member_labels: dict[int, str],
        guild_id: int,
    ) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = MEMBER_ALIAS_SHEET_NAME
        worksheet.append(list(MEMBER_ALIAS_HEADERS))
        for record in records:
            worksheet.append(
                [
                    MEMBER_ALIAS_ACTION_LABELS["keep"],
                    record.alias,
                    "",
                    member_labels.get(record.target_user_id, f"不明 ({record.target_user_id})"),
                    MEMBER_ALIAS_STATUS_LABELS.get(record.status, record.status),
                    "",
                    self._format_links(guild_id, record),
                    record.updated_at.isoformat(),
                    str(record.id),
                    record.target_user_id,
                    record.normalized_alias,
                ]
            )
            worksheet.cell(worksheet.max_row, MEMBER_ALIAS_EVIDENCE_COLUMN).value = self._format_evidences(record)
        member_sheet = workbook.create_sheet(MEMBER_ALIAS_MEMBER_SHEET_NAME)
        member_sheet.append(["対象者"])
        for label in member_labels.values():
            member_sheet.append([label])
        member_sheet.sheet_state = "hidden"
        member_range = f"'{MEMBER_ALIAS_MEMBER_SHEET_NAME}'!$A$2:$A${len(member_labels) + 1}"
        workbook.defined_names.add(DefinedName("ChatbotMemberChoices", attr_text=member_range))
        action_validation = DataValidation(type="list", formula1='"変更なし,対象者を変更,無効化"', allow_blank=False)
        member_validation = DataValidation(type="list", formula1="=ChatbotMemberChoices", allow_blank=True)
        worksheet.add_data_validation(action_validation)
        worksheet.add_data_validation(member_validation)
        action_validation.add(f"A2:A{len(records) + 1}")
        member_validation.add(f"C2:C{len(records) + 1}")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {
            "A": 18,
            "B": 24,
            "C": 36,
            "D": 36,
            "E": 12,
            "F": 72,
            "G": 48,
            "H": 28,
            "I": 38,
            "J": 22,
            "K": 24,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Meiryo UI", color="FFFFFF", bold=True)
            cell.fill = header_fill
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column != MEMBER_ALIAS_EVIDENCE_COLUMN:
                    cell.font = Font(name="Meiryo UI")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        return workbook

    def _parse_workbook(
        self,
        workbook: Workbook,
        guild: discord.Guild,
    ) -> list[MemberAliasUpdateInput]:
        worksheet = workbook[MEMBER_ALIAS_SHEET_NAME]
        headers = tuple(str(cell.value or "") for cell in worksheet[1])
        if headers != MEMBER_ALIAS_HEADERS:
            message = "列構成が出力時から変更されています"
            raise ValueError(message)
        action_values = {label: value for value, label in MEMBER_ALIAS_ACTION_LABELS.items()}
        member_values = {f"{member.display_name} ({member.id})": member.id for member in guild.members}
        updates: list[MemberAliasUpdateInput] = []
        seen_alias_ids: set[uuid.UUID] = set()
        errors: list[str] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                updates.append(self._parse_row(values, action_values, member_values, seen_alias_ids))
            except (KeyError, TypeError, ValueError) as error:
                errors.append(f"{row_number}行目: {error}")
        if errors:
            raise ValueError(" / ".join(errors[:10]))
        if not updates:
            message = "取り込み対象の行がありません"
            raise ValueError(message)
        return updates

    def _parse_row(
        self,
        values: tuple[Any, ...],
        action_values: dict[str, str],
        member_values: dict[str, int],
        seen_alias_ids: set[uuid.UUID],
    ) -> MemberAliasUpdateInput:
        action = action_values[str(values[0] or "")]
        alias = str(values[1] or "").strip()
        if not alias:
            message = "別名が空です"
            raise ValueError(message)
        target_label = str(values[2] or "")
        target_user_id = member_values.get(target_label) if action == "change_target" else None
        if action == "change_target" and target_user_id is None:
            message = "変更後の対象者が選択されていません"
            raise ValueError(message)
        alias_id = uuid.UUID(str(values[8]))
        if alias_id in seen_alias_ids:
            message = "別名IDが重複しています"
            raise ValueError(message)
        seen_alias_ids.add(alias_id)
        return MemberAliasUpdateInput(
            alias_id=alias_id,
            alias=alias,
            action=action,
            target_user_id=target_user_id,
        )

    def _format_evidences(self, record: MemberAliasReviewRecord) -> CellRichText:
        rich_text = CellRichText()
        for index, evidence in enumerate(record.evidences):
            if index:
                rich_text.append("\n")
            rich_text.append(TextBlock(InlineFont(rFont="Meiryo UI", b=True), f"{evidence.author_name}:"))
            rich_text.append(f"\n{evidence.excerpt}")
        return rich_text

    def _format_links(self, guild_id: int, record: MemberAliasReviewRecord) -> str:
        return "\n".join(
            f"https://discord.com/channels/{guild_id}/{evidence.channel_id}/{evidence.message_id}"
            for evidence in record.evidences
            if evidence.channel_id is not None
        )
