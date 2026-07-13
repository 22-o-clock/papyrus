import datetime
import io
import uuid
from logging import getLogger
from typing import Any
from zipfile import BadZipFile

import discord
from openai import AsyncOpenAI, OpenAIError
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from cogs.chatbot.constants import (
    LONG_TERM_MEMORY_EVIDENCE_COLUMN,
    LONG_TERM_MEMORY_HEADERS,
    LONG_TERM_MEMORY_KIND_LABELS,
    LONG_TERM_MEMORY_MANIFEST_SHEET_NAME,
    LONG_TERM_MEMORY_SHEET_NAME,
    LONG_TERM_MEMORY_SOURCE_LABELS,
    LONG_TERM_MEMORY_STATUS_LABELS,
    MEMBER_ALIAS_ACTION_LABELS,
    MEMBER_ALIAS_EVIDENCE_COLUMN,
    MEMBER_ALIAS_HEADERS,
    MEMBER_ALIAS_MEMBER_SHEET_NAME,
    MEMBER_ALIAS_SHEET_NAME,
    MEMBER_ALIAS_STATUS_LABELS,
    SHADOW_ACTION_LABELS,
    SHADOW_EVALUATION_FIELDS,
    SHADOW_EVALUATION_VALUES,
    SHADOW_REASON_LABELS,
    SHADOW_REVIEW_HEADERS,
)
from cogs.chatbot.observability import observe_chatbot_api_call
from cogs.chatbot.repositories.long_term_memory import (
    ChatbotLongTermMemory,
    ChatbotLongTermMemoryRepository,
    LongTermMemoryReviewRecord,
    LongTermMemoryUpdateInput,
)
from cogs.chatbot.repositories.member_alias import (
    ChatbotMemberAliasRepository,
    MemberAliasReviewRecord,
    MemberAliasUpdateInput,
)
from cogs.chatbot.repositories.shadow_candidate import (
    ChatbotShadowCandidateRepository,
    ShadowEvaluationInput,
)

from .admin_validation import (
    parse_memory_admin_expiration,
    parse_memory_admin_target,
    validate_exported_memory_ids,
)

logger = getLogger(__name__)


class ExcelManagementUseCases:
    """Chatbotのレビュー用Excelの入出力を管理します。"""

    def __init__(
        self,
        session_factory: async_sessionmaker[AsyncSession],
    ) -> None:
        self.shadow_candidate_repository = ChatbotShadowCandidateRepository(session_factory)
        self.member_alias_repository = ChatbotMemberAliasRepository(session_factory)
        self.long_term_memory_repository = ChatbotLongTermMemoryRepository(session_factory)

    async def export_chatbot_shadow_candidates(self, interaction: discord.Interaction) -> None:
        """実行した管理者が未評価の候補を最大100件、評価用Excel添付で返します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("候補の出力には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        candidates = await self.shadow_candidate_repository.get_unreviewed_candidates(interaction.user.id, limit=100)
        if not candidates:
            await interaction.response.send_message("未評価のシャドー候補はありません。", ephemeral=True)
            return
        fieldnames = [
            "trigger_message",
            "target_message",
            "conversation_context",
            "action",
            "content",
            "reaction_emoji",
            "reason",
            *SHADOW_EVALUATION_FIELDS,
            "comment",
            "created_at",
            "candidate_id",
            "channel_id",
            "trigger_message_id",
            "reply_to_message_id",
            "context_message_ids",
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "候補評価"
        worksheet.append([SHADOW_REVIEW_HEADERS[field_name] for field_name in fieldnames])
        worksheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Meiryo UI", color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for candidate in candidates:
            row = {
                "candidate_id": str(candidate.id),
                "created_at": candidate.created_at.isoformat(),
                "channel_id": candidate.channel_id,
                "trigger_message_id": candidate.trigger_message_id,
                "action": SHADOW_ACTION_LABELS.get(candidate.action, candidate.action),
                "reply_to_message_id": candidate.reply_to_message_id or "",
                "content": candidate.content,
                "reaction_emoji": candidate.reaction_emoji or "",
                "reason": SHADOW_REASON_LABELS.get(candidate.reason, candidate.reason),
                "context_message_ids": ",".join(str(message_id) for message_id in candidate.context_message_ids),
                "trigger_message": "",
                "target_message": "",
                "conversation_context": "",
                **dict.fromkeys(SHADOW_EVALUATION_FIELDS, ""),
                "comment": "",
            }
            worksheet.append([row[field] for field in fieldnames])
            context_column = fieldnames.index("conversation_context") + 1
            worksheet.cell(worksheet.max_row, context_column).value = self._format_shadow_conversation_context(
                candidate.context_snapshot
            )
            trigger_column = fieldnames.index("trigger_message") + 1
            worksheet.cell(worksheet.max_row, trigger_column).value = self._format_shadow_context_message_rich(
                candidate.context_snapshot,
                candidate.trigger_message_id,
            )
            target_column = fieldnames.index("target_message") + 1
            worksheet.cell(worksheet.max_row, target_column).value = self._format_shadow_context_message_rich(
                candidate.context_snapshot,
                candidate.reply_to_message_id,
            )
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {"A": 32, "B": 32, "C": 72, "D": 16, "E": 36, "F": 14, "G": 20, "N": 20, "O": 32}
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Meiryo UI")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        evaluation_validation = DataValidation(type="list", formula1='"◯,\u00d7,△"', allow_blank=False)
        worksheet.add_data_validation(evaluation_validation)
        for index, field_name in enumerate(fieldnames, start=1):
            if field_name in SHADOW_EVALUATION_FIELDS:
                evaluation_validation.add(
                    f"{worksheet.cell(1, index).column_letter}2:{worksheet.cell(1, index).column_letter}{len(candidates) + 1}"
                )
        worksheet.conditional_formatting.add(
            f"H2:M{len(candidates) + 1}",
            FormulaRule(formula=['H2="\u00d7"'], fill=PatternFill("solid", fgColor="F4CCCC")),
        )
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
        await interaction.response.send_message(
            f"未評価のシャドー候補 {len(candidates)} 件を出力しました。",
            ephemeral=True,
            file=discord.File(output, filename=f"chatbot_shadow_candidates_{timestamp}.xlsx"),
        )

    async def import_chatbot_shadow_evaluations(
        self,
        interaction: discord.Interaction,
        attachment: discord.Attachment,
    ) -> None:
        """Excelの有効な評価行だけを保存し、同じ管理者の既存評価を上書きします。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("評価の取込には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        workbook = load_workbook(io.BytesIO(await attachment.read()), data_only=True)
        worksheet = workbook["候補評価"]
        headers = [cell.value for cell in worksheet[1]]
        internal_headers = {display_name: field_name for field_name, display_name in SHADOW_REVIEW_HEADERS.items()}
        imported_rows = 0
        invalid_rows: list[int] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {
                internal_headers.get(str(header), str(header)): "" if value is None else str(value)
                for header, value in zip(headers, values, strict=True)
            }
            try:
                evaluation = self._parse_shadow_evaluation_row(row)
            except ValueError:
                invalid_rows.append(row_number)
                continue
            await self.shadow_candidate_repository.save_evaluation(interaction.user.id, evaluation)
            imported_rows += 1
        invalid_text = "" if not invalid_rows else f" 無効な行: {', '.join(map(str, invalid_rows))}。"
        await interaction.response.send_message(f"評価を {imported_rows} 件取り込みました。{invalid_text}", ephemeral=True)

    def _parse_shadow_evaluation_row(self, row: dict[str, str]) -> ShadowEvaluationInput:
        """CSVの1行を検証し、保存用の評価データへ変換します。"""
        if any(row.get(field) not in SHADOW_EVALUATION_VALUES for field in SHADOW_EVALUATION_FIELDS):
            raise ValueError
        try:
            candidate_id = uuid.UUID(row["candidate_id"])
        except (KeyError, ValueError) as error:
            raise ValueError from error
        return ShadowEvaluationInput(
            candidate_id=candidate_id,
            **{field: row[field] for field in SHADOW_EVALUATION_FIELDS},
            issue_category="",
            comment=row.get("comment", ""),
        )

    async def export_chatbot_member_aliases(self, interaction: discord.Interaction) -> None:
        """管理者向けに全別名と根拠を一括編集用Excelで返します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("別名の出力には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        if interaction.guild is None:
            await interaction.response.send_message("サーバー内で実行してください。", ephemeral=True)
            return
        records = await self.member_alias_repository.get_review_records()
        if not records:
            await interaction.response.send_message("保存されているメンバー別名はありません。", ephemeral=True)
            return
        member_labels = {
            member.id: f"{member.display_name} ({member.id})"
            for member in sorted(interaction.guild.members, key=lambda item: (item.display_name.casefold(), item.id))
        }
        workbook = self._build_member_alias_workbook(records, member_labels, interaction.guild.id)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
        await interaction.response.send_message(
            f"メンバー別名 {len(records)} 件を出力しました。",
            ephemeral=True,
            file=discord.File(output, filename=f"chatbot_member_aliases_{timestamp}.xlsx"),
        )

    async def import_chatbot_member_aliases(
        self,
        interaction: discord.Interaction,
        attachment: discord.Attachment,
    ) -> None:
        """全行を検証し、問題がなければ別名変更を一括適用します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("別名の取込には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        if interaction.guild is None:
            await interaction.response.send_message("サーバー内で実行してください。", ephemeral=True)
            return
        try:
            workbook = load_workbook(io.BytesIO(await attachment.read()), data_only=True)
            updates = self._parse_member_alias_workbook(workbook, interaction.guild)
            await self.member_alias_repository.apply_updates(updates, interaction.user.id)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            OpenAIError,
            SQLAlchemyError,
            TypeError,
            ValueError,
        ) as error:
            await interaction.response.send_message(f"取り込みを中止しました: {error}", ephemeral=True)
            return
        await interaction.response.send_message(f"メンバー別名 {len(updates)} 件を一括反映しました。", ephemeral=True)

    def _build_member_alias_workbook(
        self,
        records: list[MemberAliasReviewRecord],
        member_labels: dict[int, str],
        guild_id: int,
    ) -> Workbook:
        """別名一覧・根拠・入力規則を設定したExcelを作ります。"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = MEMBER_ALIAS_SHEET_NAME
        worksheet.append(list(MEMBER_ALIAS_HEADERS))
        for record in records:
            worksheet.append(
                [
                    MEMBER_ALIAS_ACTION_LABELS["keep"],
                    record.alias,
                    "",
                    member_labels.get(record.target_user_id, f"不明 ({record.target_user_id})"),
                    MEMBER_ALIAS_STATUS_LABELS.get(record.status, record.status),
                    "",
                    self._format_member_alias_links(guild_id, record),
                    record.updated_at.isoformat(),
                    str(record.id),
                    record.target_user_id,
                    record.normalized_alias,
                ]
            )
            worksheet.cell(worksheet.max_row, MEMBER_ALIAS_EVIDENCE_COLUMN).value = self._format_member_alias_evidences(record)
        member_sheet = workbook.create_sheet(MEMBER_ALIAS_MEMBER_SHEET_NAME)
        member_sheet.append(["対象者"])
        for label in member_labels.values():
            member_sheet.append([label])
        member_sheet.sheet_state = "hidden"
        member_range = f"'{MEMBER_ALIAS_MEMBER_SHEET_NAME}'!$A$2:$A${len(member_labels) + 1}"
        workbook.defined_names.add(DefinedName("ChatbotMemberChoices", attr_text=member_range))
        action_validation = DataValidation(type="list", formula1='"変更なし,対象者を変更,無効化"', allow_blank=False)
        member_validation = DataValidation(type="list", formula1="=ChatbotMemberChoices", allow_blank=True)
        worksheet.add_data_validation(action_validation)
        worksheet.add_data_validation(member_validation)
        action_validation.add(f"A2:A{len(records) + 1}")
        member_validation.add(f"C2:C{len(records) + 1}")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {"A": 18, "B": 24, "C": 36, "D": 36, "E": 12, "F": 72, "G": 48, "H": 28, "I": 38, "J": 22, "K": 24}
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Meiryo UI", color="FFFFFF", bold=True)
            cell.fill = header_fill
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column != MEMBER_ALIAS_EVIDENCE_COLUMN:
                    cell.font = Font(name="Meiryo UI")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        return workbook

    def _parse_member_alias_workbook(
        self,
        workbook: Workbook,
        guild: discord.Guild,
    ) -> list[MemberAliasUpdateInput]:
        """別名Excelの全行を検証し、DB更新用データへ変換します。"""
        worksheet = workbook[MEMBER_ALIAS_SHEET_NAME]
        headers = tuple(str(cell.value or "") for cell in worksheet[1])
        if headers != MEMBER_ALIAS_HEADERS:
            msg = "列構成が出力時から変更されています"
            raise ValueError(msg)
        action_values = {label: value for value, label in MEMBER_ALIAS_ACTION_LABELS.items()}
        member_values = {f"{member.display_name} ({member.id})": member.id for member in guild.members}
        updates: list[MemberAliasUpdateInput] = []
        seen_alias_ids: set[uuid.UUID] = set()
        errors: list[str] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                updates.append(
                    self._parse_member_alias_row(
                        values,
                        action_values,
                        member_values,
                        seen_alias_ids,
                    )
                )
            except (KeyError, TypeError, ValueError) as error:
                errors.append(f"{row_number}行目: {error}")
        if errors:
            raise ValueError(" / ".join(errors[:10]))
        if not updates:
            msg = "取り込み対象の行がありません"
            raise ValueError(msg)
        return updates

    def _parse_member_alias_row(
        self,
        values: tuple[Any, ...],
        action_values: dict[str, str],
        member_values: dict[str, int],
        seen_alias_ids: set[uuid.UUID],
    ) -> MemberAliasUpdateInput:
        """別名Excelの1行を検証します。"""
        action = action_values[str(values[0] or "")]
        alias = str(values[1] or "").strip()
        if not alias:
            msg = "別名が空です"
            raise ValueError(msg)
        target_label = str(values[2] or "")
        target_user_id = member_values.get(target_label) if action == "change_target" else None
        if action == "change_target" and target_user_id is None:
            msg = "変更後の対象者が選択されていません"
            raise ValueError(msg)
        alias_id = uuid.UUID(str(values[8]))
        if alias_id in seen_alias_ids:
            msg = "別名IDが重複しています"
            raise ValueError(msg)
        seen_alias_ids.add(alias_id)
        return MemberAliasUpdateInput(
            alias_id=alias_id,
            alias=alias,
            action=action,
            target_user_id=target_user_id,
        )

    def _format_member_alias_evidences(self, record: MemberAliasReviewRecord) -> CellRichText:
        """根拠投稿の発言者だけを太字にしたExcelセル値を作ります。"""
        rich_text = CellRichText()
        for index, evidence in enumerate(record.evidences):
            if index:
                rich_text.append("\n")
            rich_text.append(TextBlock(InlineFont(rFont="Meiryo UI", b=True), f"{evidence.author_name}:"))
            rich_text.append(f"\n{evidence.excerpt}")
        return rich_text

    def _format_member_alias_links(self, guild_id: int, record: MemberAliasReviewRecord) -> str:
        """根拠投稿のDiscordリンクを改行区切りで返します。"""
        return "\n".join(
            f"https://discord.com/channels/{guild_id}/{evidence.channel_id}/{evidence.message_id}"
            for evidence in record.evidences
            if evidence.channel_id is not None
        )

    async def export_chatbot_memories(self, interaction: discord.Interaction) -> None:
        """管理者向けに全長期記憶を一括編集用Excelで返します。"""
        if not interaction.permissions.manage_guild or interaction.guild is None:
            await interaction.response.send_message("長期記憶の出力には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        records = await self.long_term_memory_repository.get_review_records()
        workbook = self._build_long_term_memory_workbook(records, interaction.guild)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
        await interaction.response.send_message(
            f"長期記憶 {len(records)} 件を出力しました。",
            ephemeral=True,
            file=discord.File(output, filename=f"chatbot_memories_{timestamp}.xlsx"),
        )

    async def import_chatbot_memories(
        self,
        interaction: discord.Interaction,
        attachment: discord.Attachment,
    ) -> None:
        """全行を検証し、必要な埋め込みを生成してから長期記憶を一括更新します。"""
        if not interaction.permissions.manage_guild or interaction.guild is None:
            await interaction.response.send_message("長期記憶の取込には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True)
        try:
            workbook = load_workbook(io.BytesIO(await attachment.read()), data_only=True)
            records = await self.long_term_memory_repository.get_review_records()
            updates = self._parse_long_term_memory_workbook(workbook, interaction.guild, records)
            await self._create_updated_memory_embeddings(updates, records)
            changed_count = await self.long_term_memory_repository.apply_admin_updates(updates, interaction.user.id)
            logger.info(
                "Applied chatbot memory admin updates (administrator_user_id=%s, changed_count=%s, checked_count=%s)",
                interaction.user.id,
                changed_count,
                len(updates),
            )
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            OpenAIError,
            SQLAlchemyError,
            TypeError,
            ValueError,
        ) as error:
            await interaction.followup.send(f"取り込みを中止しました: {error}", ephemeral=True)
            return
        await interaction.followup.send(
            f"長期記憶を {changed_count} 件変更しました (確認 {len(updates)} 件)。",
            ephemeral=True,
        )

    def _build_long_term_memory_workbook(
        self,
        records: list[LongTermMemoryReviewRecord],
        guild: discord.Guild,
    ) -> Workbook:
        """全長期記憶と編集用入力規則を設定したExcelを作ります。"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = LONG_TERM_MEMORY_SHEET_NAME
        worksheet.append(list(LONG_TERM_MEMORY_HEADERS))
        member_labels = {
            member.id: f"{member.display_name} ({member.id})"
            for member in sorted(guild.members, key=lambda item: (item.display_name.casefold(), item.id))
        }
        now = datetime.datetime.now(datetime.UTC)
        for record in records:
            memory = record.memory
            current_target = (
                member_labels.get(memory.target_user_id, f"不明 ({memory.target_user_id})")
                if memory.target_user_id is not None
                else memory.external_entity_name or "共有情報"
            )
            editable_target = (
                member_labels.get(memory.target_user_id, "")
                if memory.target_user_id is not None
                else memory.external_entity_name or ""
            )
            status_key = (
                "expired" if memory.status == "active" and memory.expires_at and memory.expires_at <= now else memory.status
            )
            status = LONG_TERM_MEMORY_STATUS_LABELS.get(status_key, status_key)
            worksheet.append(
                [
                    "変更なし",
                    memory.content,
                    self._memory_target_type_label(memory.target_resolution),
                    editable_target,
                    current_target,
                    LONG_TERM_MEMORY_KIND_LABELS.get(memory.kind, memory.kind),
                    LONG_TERM_MEMORY_SOURCE_LABELS.get(memory.source_type, memory.source_type),
                    "はい" if memory.is_sensitive else "いいえ",
                    self._memory_datetime_for_excel(memory.expires_at),
                    status,
                    "",
                    self._format_long_term_memory_links(guild.id, record),
                    self._memory_datetime_for_excel(memory.observed_at),
                    self._memory_datetime_for_excel(memory.created_at),
                    str(memory.superseded_by_memory_id or ""),
                    str(memory.conflict_group_id or ""),
                    str(memory.id),
                ]
            )
            worksheet.cell(worksheet.max_row, 11).value = self._format_long_term_memory_evidences(record)
        manifest_sheet = workbook.create_sheet(LONG_TERM_MEMORY_MANIFEST_SHEET_NAME)
        manifest_sheet.append(["記憶ID"])
        for record in records:
            manifest_sheet.append([str(record.memory.id)])
        manifest_sheet.sheet_state = "hidden"
        self._add_long_term_memory_validations(workbook, worksheet, member_labels, len(records))
        self._style_long_term_memory_worksheet(worksheet)
        return workbook

    def _add_long_term_memory_validations(
        self,
        workbook: Workbook,
        worksheet: Worksheet,
        member_labels: dict[int, str],
        record_count: int,
    ) -> None:
        """管理入力列へプルダウン候補を設定します。"""
        member_sheet = workbook.create_sheet(MEMBER_ALIAS_MEMBER_SHEET_NAME)
        member_sheet.append(["対象者"])
        for label in member_labels.values():
            member_sheet.append([label])
        member_sheet.sheet_state = "hidden"
        member_range = f"'{MEMBER_ALIAS_MEMBER_SHEET_NAME}'!$A$2:$A${len(member_labels) + 1}"
        workbook.defined_names.add(DefinedName("ChatbotMemoryMemberChoices", attr_text=member_range))
        validations = (
            (DataValidation(type="list", formula1='"変更なし,更新,無効化,有効化"'), f"A2:A{record_count + 1}"),
            (DataValidation(type="list", formula1='"メンバー,外部対象,共有情報"'), f"C2:C{record_count + 1}"),
            (
                DataValidation(type="list", formula1="=ChatbotMemoryMemberChoices", showErrorMessage=False),
                f"D2:D{record_count + 1}",
            ),
            (DataValidation(type="list", formula1='"プロフィール,継続中,一時的,共有"'), f"F2:F{record_count + 1}"),
            (DataValidation(type="list", formula1='"本人発言,第三者発言,推測"'), f"G2:G{record_count + 1}"),
            (DataValidation(type="list", formula1='"はい,いいえ"'), f"H2:H{record_count + 1}"),
        )
        for validation, cell_range in validations:
            worksheet.add_data_validation(validation)
            validation.add(cell_range)

    def _style_long_term_memory_worksheet(self, worksheet: Worksheet) -> None:
        """長期記憶ExcelをPC上で確認しやすい幅と書式へ整えます。"""
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {
            "A": 16,
            "B": 48,
            "C": 14,
            "D": 36,
            "E": 36,
            "F": 14,
            "G": 14,
            "H": 12,
            "I": 22,
            "J": 14,
            "K": 72,
            "L": 48,
            "M": 22,
            "N": 22,
            "O": 38,
            "P": 38,
            "Q": 38,
        }
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Meiryo UI", color="FFFFFF", bold=True)
            cell.fill = header_fill
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column != LONG_TERM_MEMORY_EVIDENCE_COLUMN:
                    cell.font = Font(name="Meiryo UI")
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _parse_long_term_memory_workbook(
        self,
        workbook: Workbook,
        guild: discord.Guild,
        records: list[LongTermMemoryReviewRecord],
    ) -> list[LongTermMemoryUpdateInput]:
        """長期記憶Excelを全行検証し、更新データへ変換します。"""
        worksheet = workbook[LONG_TERM_MEMORY_SHEET_NAME]
        headers = tuple(str(cell.value or "") for cell in worksheet[1])
        if headers != LONG_TERM_MEMORY_HEADERS:
            msg = "列構成が出力時から変更されています"
            raise ValueError(msg)
        memories = {record.memory.id: record.memory for record in records}
        member_values = {f"{member.display_name} ({member.id})": member.id for member in guild.members}
        updates: list[LongTermMemoryUpdateInput] = []
        seen_ids: set[uuid.UUID] = set()
        errors: list[str] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                update = self._parse_long_term_memory_row(values, member_values)
                self._validate_memory_update_id(update.memory_id, memories, seen_ids)
                seen_ids.add(update.memory_id)
                updates.append(update)
            except (KeyError, TypeError, ValueError) as error:
                errors.append(f"{row_number}行目: {error}")
        if errors:
            raise ValueError(" / ".join(errors[:10]))
        self._validate_memory_manifest(workbook, seen_ids)
        return updates

    def _validate_memory_manifest(self, workbook: Workbook, seen_ids: set[uuid.UUID]) -> None:
        """新形式のExcelでは、出力後に行が削除されていないことを確認します。"""
        if LONG_TERM_MEMORY_MANIFEST_SHEET_NAME not in workbook.sheetnames:
            return
        worksheet = workbook[LONG_TERM_MEMORY_MANIFEST_SHEET_NAME]
        exported_ids = {
            uuid.UUID(str(value))
            for (value,) in worksheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if value not in (None, "")
        }
        validate_exported_memory_ids(seen_ids, exported_ids)

    def _validate_memory_update_id(
        self,
        memory_id: uuid.UUID,
        memories: dict[uuid.UUID, ChatbotLongTermMemory],
        seen_ids: set[uuid.UUID],
    ) -> None:
        """記憶IDが出力対象に存在し、Excel内で一意であることを保証します。"""
        if memory_id not in memories or memory_id in seen_ids:
            msg = "記憶IDが存在しないか重複しています"
            raise ValueError(msg)

    def _parse_long_term_memory_row(
        self,
        values: tuple[Any, ...],
        member_values: dict[str, int],
    ) -> LongTermMemoryUpdateInput:
        """長期記憶Excelの1行を検証します。"""
        actions = {"変更なし": "keep", "更新": "update", "無効化": "invalidate", "有効化": "activate"}
        kinds = {label: value for value, label in LONG_TERM_MEMORY_KIND_LABELS.items()}
        sources = {label: value for value, label in LONG_TERM_MEMORY_SOURCE_LABELS.items()}
        action = actions[str(values[0] or "")]
        content = str(values[1] or "").strip()
        if not content:
            msg = "内容が空です"
            raise ValueError(msg)
        target_user_id, external_name, resolution = parse_memory_admin_target(
            str(values[2] or ""), str(values[3] or ""), member_values
        )
        sensitive_text = str(values[7] or "")
        if sensitive_text not in {"はい", "いいえ"}:
            msg = "機微情報は「はい」か「いいえ」で指定してください"
            raise ValueError(msg)
        return LongTermMemoryUpdateInput(
            memory_id=uuid.UUID(str(values[16])),
            action=action,
            content=content,
            target_user_id=target_user_id,
            external_entity_name=external_name,
            target_resolution=resolution,
            kind=kinds[str(values[5] or "")],
            source_type=sources[str(values[6] or "")],
            is_sensitive=sensitive_text == "はい",
            expires_at=parse_memory_admin_expiration(values[8]),
            embedding=None,
        )

    def _memory_datetime_for_excel(self, value: datetime.datetime | None) -> datetime.datetime | None:
        """タイムゾーン付き日時をExcel表示用の日本時間へ変換します。"""
        if value is None:
            return None
        japan_timezone = datetime.timezone(datetime.timedelta(hours=9))
        return value.astimezone(japan_timezone).replace(tzinfo=None)

    async def _create_updated_memory_embeddings(
        self,
        updates: list[LongTermMemoryUpdateInput],
        records: list[LongTermMemoryReviewRecord],
    ) -> None:
        """内容が変更された記憶だけの埋め込みを一括生成します。"""
        current_contents = {record.memory.id: record.memory.content for record in records}
        changed = [item for item in updates if item.content != current_contents[item.memory_id]]
        if not changed:
            return
        response = await observe_chatbot_api_call(
            "memory_admin_embedding",
            "text-embedding-3-large",
            AsyncOpenAI().embeddings.create(model="text-embedding-3-large", input=[item.content for item in changed]),
            item_count=len(changed),
        )
        for item, embedding_data in zip(changed, response.data, strict=True):
            item.embedding = embedding_data.embedding

    def _memory_target_type_label(self, resolution: str) -> str:
        """DBの対象解決状態をExcel表示へ変換します。"""
        return {"member": "メンバー", "external": "外部対象"}.get(resolution, "共有情報")

    def _format_long_term_memory_evidences(self, record: LongTermMemoryReviewRecord) -> CellRichText:
        """根拠投稿の発言者だけを太字にしたExcelセル値を作ります。"""
        result = CellRichText()
        for index, evidence in enumerate(record.evidences):
            if index:
                result.append("\n")
            result.append(TextBlock(InlineFont(rFont="Meiryo UI", b=True), f"{evidence.author_name}:"))
            result.append(f"\n{evidence.excerpt}")
        return result

    def _format_long_term_memory_links(self, guild_id: int, record: LongTermMemoryReviewRecord) -> str:
        """根拠投稿のDiscordリンクを改行区切りで返します。"""
        return "\n".join(
            f"https://discord.com/channels/{guild_id}/{evidence.channel_id}/{evidence.message_id}"
            for evidence in record.evidences
            if evidence.channel_id is not None
        )

    def _format_shadow_context_message(
        self,
        context_snapshot: list[dict[str, object]],
        message_id: int | None,
    ) -> str:
        """保存済みの文脈から、CSV表示用の発言を整形します。"""
        if message_id is None:
            return ""
        for message in context_snapshot:
            if message["message_id"] == message_id:
                return f"{message['author_name']}: {message['content']}"
        return ""

    def _format_shadow_context_message_rich(
        self,
        context_snapshot: list[dict[str, object]],
        message_id: int | None,
    ) -> CellRichText:
        """反応元または反応対象を、発言者名を太字にしたExcel用リッチテキストへ変換します。"""
        if message_id is None:
            return CellRichText()
        for message in context_snapshot:
            if message["message_id"] == message_id:
                return CellRichText(
                    TextBlock(InlineFont(rFont="Meiryo UI", b=True), f"{message['author_name']}: "),
                    TextBlock(InlineFont(rFont="Meiryo UI"), str(message["content"])),
                )
        return CellRichText()

    def _format_shadow_conversation_context(self, context_snapshot: list[dict[str, object]]) -> CellRichText:
        """会話抜粋を、発言者名を太字にしたExcel用リッチテキストへ変換します。"""
        context = CellRichText()
        for index, message in enumerate(context_snapshot):
            author_font = InlineFont(rFont="Meiryo UI", b=True)
            content_font = InlineFont(rFont="Meiryo UI")
            context.append(TextBlock(author_font, f"{message['author_name']}: "))
            suffix = "\n" if index < len(context_snapshot) - 1 else ""
            context.append(TextBlock(content_font, f"{message['content']}{suffix}"))
        return context
