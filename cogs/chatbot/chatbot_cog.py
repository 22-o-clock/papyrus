import asyncio
import datetime
import io
import json
import random
import re
import uuid
from dataclasses import dataclass, field
from logging import getLogger
from typing import Any, cast

import discord
from discord import Message, MessageReference, app_commands
from discord.ext import commands
from openai import AsyncOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from .channel_roles import ChannelRole, ChannelRoleManager
from .database import (
    ChatbotLongTermMemory,
    ChatbotLongTermMemoryStore,
    ChatbotMemberAliasStore,
    ChatbotMemoryExtractionQueueStore,
    ChatbotShadowCandidateStore,
    ChatbotShortTermMessageStore,
    LongTermMemoryInput,
    MemberAliasInput,
    ShadowCandidateInput,
    ShadowEvaluationInput,
    StoredAttachmentInput,
    StoredMessageInput,
    find_user_ids_by_member_alias,
    normalize_member_alias,
)
from .database_envs import DatabaseEnvManager
from .responses_api import (
    AttachmentInMemory,
    LLMMessage,
    LongTermMemoryExtractor,
    MessageInMemory,
    ResponseAction,
    ResponsePipeline,
    ShadowReason,
)
from .shadow_mode import ShadowModeManager

logger = getLogger(__name__)

ASSISTANT_DEBOUNCE_SECONDS = 2.0
CHAT_DEBOUNCE_MIN_SECONDS = 5.0
CHAT_DEBOUNCE_MAX_SECONDS = 15.0
CHAT_TEXT_COOLDOWN_SECONDS = 15 * 60
CHAT_REACTION_COOLDOWN_SECONDS = 2 * 60
DEFAULT_CONVERSATION_RESET_MINUTES = 12 * 60
MINIMUM_CONVERSATION_RESET_MINUTES = 1
CONVERSATION_RESET_MINUTES_KEY = "CHATBOT_CONVERSATION_RESET_MINUTES"
DEFAULT_UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES = 30
DEFAULT_UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES = 60
UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES_KEY = "CHATBOT_UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES"
UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES_KEY = "CHATBOT_UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES"
ATTACHMENT_CONTEXT_MAX_CHARACTERS = 100
MEMORY_EXTRACTION_BATCH_SIZE = 5
MEMORY_EXTRACTION_WAIT_SECONDS = 10 * 60
MEMORY_SEARCH_CONTEXT_MESSAGE_COUNT = 10
MEMORY_SEARCH_MAXIMUM_COSINE_DISTANCE = 0.70

QUESTION_ENDING_PATTERN = re.compile(r"(?:\?|ですか|ますか|でしょうか|かな|の\?|何\?|どう\?|誰\?|どこ\?|いつ\?)$")
SHADOW_EVALUATION_FIELDS = (
    "action_appropriate",
    "context_understood",
    "identity_correct",
    "length_natural",
    "non_intrusive",
    "worth_posting",
)
SHADOW_EVALUATION_VALUES = {"◯", "\u00d7", "△"}
SHADOW_REVIEW_HEADERS = {
    "trigger_message": "反応元メッセージ",
    "target_message": "反応対象メッセージ",
    "conversation_context": "会話抜粋",
    "action": "選択した行動",
    "content": "生成文",
    "reaction_emoji": "リアクション",
    "reason": "判断理由",
    "action_appropriate": "行動選択の適切さ",
    "context_understood": "文脈の理解",
    "identity_correct": "人物の区別",
    "length_natural": "長さの自然さ",
    "non_intrusive": "邪魔でない",
    "worth_posting": "総合評価",
    "comment": "コメント",
    "created_at": "作成日時",
    "candidate_id": "候補ID",
    "channel_id": "チャンネルID",
    "trigger_message_id": "反応元メッセージID",
    "reply_to_message_id": "反応対象メッセージID",
    "context_message_ids": "文脈メッセージID一覧",
}
SHADOW_ACTION_LABELS = {
    "silence": "沈黙",
    "reaction": "リアクション",
    "reply": "返信",
    "message": "通常投稿",
}
SHADOW_REASON_LABELS = {
    "natural_contribution": "自然な会話",
    "helpful_unanswered_question": "未回答質問への回答",
    "avoid_interrupting_humans": "人間の会話を優先",
    "no_helpful_contribution": "有益な回答ができない",
    "identity_uncertain": "発言者を区別できない",
    "cooldown": "クールダウン中",
}


class AttachmentAnalysis(BaseModel):
    """短期文脈に保存する添付ファイルの要約です。"""

    summary: str
    important_text: str


@dataclass
class ChannelProcessingState:
    """チャンネルごとの生成状態と生成中に受信したメッセージを保持します。"""

    lock: asyncio.Lock = field(default_factory=asyncio.Lock)
    generating: bool = False
    pending_messages: list[Message] = field(default_factory=list)
    queued_response_message: Message | None = None
    queued_response_is_explicit_call: bool = False
    debounce_task: asyncio.Task[None] | None = None
    debounced_response_message: Message | None = None
    debounced_response_is_explicit_call: bool = False
    generation_revision: int = 0
    last_spontaneous_action_at: float | None = None
    last_human_message_timestamp: datetime.datetime | None = None
    unanswered_question_task: asyncio.Task[None] | None = None
    unanswered_question_message_id: int | None = None
    queued_response_is_unanswered_question: bool = False
    debounced_response_is_unanswered_question: bool = False


def claim_response_slot(
    state: ChannelProcessingState,
    message: Message,
    *,
    is_explicit_call: bool,
    is_unanswered_question: bool,
) -> bool:
    """生成枠を確保し、使用中の場合は次の返信対象としてメッセージを保持します。"""
    if state.generating:
        state.queued_response_message = message
        state.queued_response_is_explicit_call = is_explicit_call
        state.queued_response_is_unanswered_question = is_unanswered_question
        state.generation_revision += 1
        return False

    state.generating = True
    return True


def get_response_debounce_seconds(role: ChannelRole) -> float:
    """役割に応じた返信生成前の待機秒数を返します。"""
    if role is ChannelRole.ASSISTANT:
        return ASSISTANT_DEBOUNCE_SECONDS
    return random.SystemRandom().uniform(CHAT_DEBOUNCE_MIN_SECONDS, CHAT_DEBOUNCE_MAX_SECONDS)


def is_generation_current(state: ChannelProcessingState, revision: int) -> bool:
    """生成開始後に、回答を作り直す必要がある返信要求が追加されていないか確認します。"""
    return state.generation_revision == revision


def can_execute_spontaneous_action(
    action: ResponseAction,
    last_action_at: float | None,
    now: float,
) -> bool:
    """自発反応が行動別のクールダウンを過ぎているか判定します。"""
    if action is ResponseAction.SILENCE or last_action_at is None:
        return True

    cooldown_seconds = CHAT_REACTION_COOLDOWN_SECONDS if action is ResponseAction.REACTION else CHAT_TEXT_COOLDOWN_SECONDS
    return now - last_action_at >= cooldown_seconds


def can_start_spontaneous_generation(last_action_at: float | None, now: float) -> bool:
    """全ての自発行動が抑制される期間を避けて生成を始めるか判定します。"""
    if last_action_at is None:
        return True
    return now - last_action_at >= CHAT_REACTION_COOLDOWN_SECONDS


def should_reset_conversation(
    last_human_message_timestamp: datetime.datetime | None,
    current_message_timestamp: datetime.datetime,
    reset_minutes: int,
) -> bool:
    """最後の人間投稿から設定時間以上空いたときに会話文脈をリセットするか判定します。"""
    if last_human_message_timestamp is None:
        return False
    return current_message_timestamp - last_human_message_timestamp >= datetime.timedelta(minutes=reset_minutes)


def is_unaddressed_question(
    *,
    content: str,
    is_reply: bool,
    mentioned_user_ids: list[int],
) -> bool:
    """宛先のない質問として待機対象にする投稿か判定します。"""
    if is_reply or mentioned_user_ids:
        return False
    normalized_content = content.replace("\uff1f", "?").strip()
    return QUESTION_ENDING_PATTERN.search(normalized_content) is not None


def get_unanswered_question_wait_minutes(minimum_minutes: int, maximum_minutes: int) -> int:
    """宛先のない質問への回答を待つ時間を一様ランダムに選びます。"""
    return random.SystemRandom().randint(minimum_minutes, maximum_minutes)


def can_change_channel_role(*, is_thread: bool, manage_channels: bool) -> bool:
    """スレッドでは全員、通常チャンネルでは管理権限を持つ人だけに変更を許可します。"""
    return is_thread or manage_channels


def get_available_referenced_author_id(reference: MessageReference) -> int | None:
    """追加のAPI取得なしで利用できる返信元メッセージの発言者IDを返します。"""
    if isinstance(reference.resolved, Message):
        return reference.resolved.author.id
    if reference.cached_message is not None:
        return reference.cached_message.author.id
    return None


def should_respond(
    role: ChannelRole,
    *,
    mentioned_bot: bool,
    replied_to_bot: bool,
) -> bool:
    """チャンネル役割と呼びかけ方法から、応答判断を開始するか決定します。"""
    explicitly_called = mentioned_bot or replied_to_bot
    return explicitly_called or role is ChannelRole.CHAT


def get_latest_memory_search_query(message: MessageInMemory) -> str:
    """最新投稿は識別用メタデータを除き、本文の意味を優先して記憶検索へ使います。"""
    content = message.content.strip()
    return content or json.dumps(message.to_dict(), ensure_ascii=False)


class ChatBot(commands.Cog):
    def __init__(self, bot: commands.Bot, session_factory: async_sessionmaker[AsyncSession]) -> None:
        self.bot = bot
        self.response_pipelines: dict[int, ResponsePipeline] = {}
        self.env_manager = DatabaseEnvManager(session_factory)
        self.channel_role_manager = ChannelRoleManager(self.env_manager)
        self.shadow_mode_manager = ShadowModeManager(self.env_manager)
        self.shadow_candidate_store = ChatbotShadowCandidateStore(session_factory)
        self.short_term_message_store = ChatbotShortTermMessageStore(session_factory)
        self.memory_extraction_queue = ChatbotMemoryExtractionQueueStore(session_factory)
        self.long_term_memory_store = ChatbotLongTermMemoryStore(session_factory)
        self.member_alias_store = ChatbotMemberAliasStore(session_factory)
        self.long_term_memory_extractor = LongTermMemoryExtractor(AsyncOpenAI())
        self._memory_extraction_task: asyncio.Task[None] | None = None
        self._memory_queue_recovered = False

        self._initialization_lock = asyncio.Lock()
        self._channel_states: dict[int, ChannelProcessingState] = {}
        self._background_tasks: set[asyncio.Task[None]] = set()
        self.conversation_reset_minutes = DEFAULT_CONVERSATION_RESET_MINUTES
        self.unanswered_question_minimum_wait_minutes = DEFAULT_UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES
        self.unanswered_question_maximum_wait_minutes = DEFAULT_UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES

    async def initialize_response_pipeline_for_channel(self, channel_id: int) -> None:
        if self.bot.user:
            self.response_pipelines[channel_id] = ResponsePipeline(AsyncOpenAI(), self.bot.user.display_name)
        else:
            logger.warning(
                "Bot user is not available during initialization, response pipeline may not be initialized correctly"
            )
            self.response_pipelines[channel_id] = ResponsePipeline(AsyncOpenAI(), "Bot")

    async def _ensure_channel_state(self, channel_id: int) -> ChannelProcessingState:
        """チャンネルの応答パイプラインと処理状態を一度だけ初期化します。"""
        if channel_id not in self._channel_states:
            async with self._initialization_lock:
                if channel_id not in self._channel_states:
                    await self.initialize_response_pipeline_for_channel(channel_id)
                    stored_messages = await self.short_term_message_store.get_for_channel(channel_id)
                    stored_attachments = await self.short_term_message_store.get_attachments(
                        [stored.message_id for stored in stored_messages]
                    )
                    attachments_by_message_id: dict[int, list[AttachmentInMemory]] = {}
                    for attachment in stored_attachments:
                        attachments_by_message_id.setdefault(attachment.message_id, []).append(
                            AttachmentInMemory(
                                attachment_id=attachment.id,
                                filename=attachment.filename,
                                kind=attachment.kind,
                                analysis_status=attachment.analysis_status,
                                summary=self._truncate_attachment_context(attachment.summary),
                                important_text=self._truncate_attachment_context(attachment.important_text),
                            )
                        )
                    self.response_pipelines[channel_id].short_term_memory.restore(
                        [
                            MessageInMemory(
                                message_id=stored.message_id,
                                author_id=stored.author_id,
                                author_name=stored.author_name,
                                content=stored.content,
                                reply_to_message_id=stored.reply_to_message_id,
                                mentioned_user_ids=stored.mentioned_user_ids,
                                timestamp=stored.created_at,
                                attachments=attachments_by_message_id.get(stored.message_id, []),
                            )
                            for stored in stored_messages
                        ]
                    )
                    last_human_message = next(
                        (stored for stored in reversed(stored_messages) if not stored.is_bot),
                        None,
                    )
                    self._channel_states[channel_id] = ChannelProcessingState(
                        last_human_message_timestamp=(last_human_message.created_at if last_human_message is not None else None)
                    )
        return self._channel_states[channel_id]

    @commands.Cog.listener()
    async def on_ready(self) -> None:
        """サーバー共通の待機時間設定を読み込みます。"""
        self.conversation_reset_minutes = await self._load_positive_minutes(
            CONVERSATION_RESET_MINUTES_KEY,
            DEFAULT_CONVERSATION_RESET_MINUTES,
        )
        minimum_minutes = await self._load_positive_minutes(
            UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES_KEY,
            DEFAULT_UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES,
        )
        maximum_minutes = await self._load_positive_minutes(
            UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES_KEY,
            DEFAULT_UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES,
        )
        if minimum_minutes > maximum_minutes:
            logger.warning(
                "Invalid chatbot unanswered question wait range (minimum=%s, maximum=%s)",
                minimum_minutes,
                maximum_minutes,
            )
            return
        self.unanswered_question_minimum_wait_minutes = minimum_minutes
        self.unanswered_question_maximum_wait_minutes = maximum_minutes
        if not self._memory_queue_recovered:
            await self.memory_extraction_queue.recover_interrupted()
            self._memory_queue_recovered = True
        await self._schedule_memory_extraction()

    async def _load_positive_minutes(self, key: str, default: int) -> int:
        """DBに保存した分単位の正の整数設定を読み込み、異常値は既定値へ戻します。"""
        configured_minutes = await self.env_manager.get_env(key)
        if configured_minutes is None:
            return default

        try:
            minutes = int(configured_minutes)
        except ValueError:
            logger.warning("Invalid chatbot minutes setting (key=%s, value=%r)", key, configured_minutes)
            return default

        if minutes < MINIMUM_CONVERSATION_RESET_MINUTES:
            logger.warning("Chatbot minutes setting is too small (key=%s, value=%s)", key, minutes)
            return default
        return minutes

    @app_commands.command(name="show_chatbot_role", description="このチャンネルでのChatbotの役割を表示します")
    async def show_chatbot_role(self, interaction: discord.Interaction) -> None:
        channel_id = interaction.channel_id
        if channel_id is None:
            await interaction.response.send_message("チャンネル情報を取得できませんでした。", ephemeral=True)
            return

        is_thread = isinstance(interaction.channel, discord.Thread)
        parent_channel_id = interaction.channel.parent_id if is_thread else None
        configured_role = await self.channel_role_manager.get_configured_role(channel_id)
        role = await self.channel_role_manager.get_role(channel_id, parent_channel_id)
        source = "このスレッド固有の設定" if configured_role is not None else "親チャンネルからの継承"
        if not is_thread:
            source = "このチャンネルの設定" if configured_role is not None else "既定値"
        await interaction.response.send_message(
            f"このチャンネルのChatbotの役割は `{role.value}` です。設定元: {source}。",
            ephemeral=True,
        )

    @app_commands.command(name="set_chatbot_reset_minutes", description="Chatbotの会話リセット時間を変更します")
    @app_commands.describe(minutes="最後の人間投稿から会話をリセットするまでの分数 (1以上)")
    async def set_chatbot_conversation_reset_minutes(
        self,
        interaction: discord.Interaction,
        minutes: int,
    ) -> None:
        """サーバー全体の会話リセット時間を保存します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message(
                "会話リセット時間の変更には「サーバー管理」権限が必要です。",
                ephemeral=True,
            )
            return
        if minutes < MINIMUM_CONVERSATION_RESET_MINUTES:
            await interaction.response.send_message("会話リセット時間は1分以上で指定してください。", ephemeral=True)
            return

        previous_minutes = self.conversation_reset_minutes
        self.conversation_reset_minutes = minutes
        await self.env_manager.set_env(CONVERSATION_RESET_MINUTES_KEY, str(minutes))
        await interaction.response.send_message(
            f"Chatbotの会話リセット時間を {previous_minutes}分から {minutes}分に変更しました。"
        )

    @app_commands.command(name="set_chatbot_question_wait", description="宛先のない質問への回答待機時間を変更します")
    @app_commands.describe(
        minimum_minutes="最短待機時間 (分、1以上)",
        maximum_minutes="最長待機時間 (分、最短以上)",
    )
    async def set_chatbot_question_wait(
        self,
        interaction: discord.Interaction,
        minimum_minutes: int,
        maximum_minutes: int,
    ) -> None:
        """サーバー全体の宛先のない質問への待機時間を保存します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message(
                "質問への回答待機時間の変更には「サーバー管理」権限が必要です。",
                ephemeral=True,
            )
            return
        if minimum_minutes < MINIMUM_CONVERSATION_RESET_MINUTES or minimum_minutes > maximum_minutes:
            await interaction.response.send_message(
                "待機時間は「1以上の最短分数」と「最短以上の最長分数」で指定してください。",
                ephemeral=True,
            )
            return

        previous_minimum_minutes = self.unanswered_question_minimum_wait_minutes
        previous_maximum_minutes = self.unanswered_question_maximum_wait_minutes
        self.unanswered_question_minimum_wait_minutes = minimum_minutes
        self.unanswered_question_maximum_wait_minutes = maximum_minutes
        await self.env_manager.set_env(UNANSWERED_QUESTION_MINIMUM_WAIT_MINUTES_KEY, str(minimum_minutes))
        await self.env_manager.set_env(UNANSWERED_QUESTION_MAXIMUM_WAIT_MINUTES_KEY, str(maximum_minutes))
        await interaction.response.send_message(
            "宛先のない質問への回答待機時間を "
            f"{previous_minimum_minutes}〜{previous_maximum_minutes}分から "
            f"{minimum_minutes}〜{maximum_minutes}分に変更しました。"
        )

    @app_commands.command(name="set_chatbot_role", description="このチャンネルでのChatbotの役割を変更します")
    @app_commands.describe(role="assistant または chat を選択します")
    async def set_chatbot_role(self, interaction: discord.Interaction, role: ChannelRole) -> None:
        channel_id = interaction.channel_id
        if channel_id is None:
            await interaction.response.send_message("チャンネル情報を取得できませんでした。", ephemeral=True)
            return

        is_thread = isinstance(interaction.channel, discord.Thread)
        if not can_change_channel_role(
            is_thread=is_thread,
            manage_channels=interaction.permissions.manage_channels,
        ):
            await interaction.response.send_message(
                "通常チャンネルの役割変更には「チャンネルの管理」権限が必要です。",
                ephemeral=True,
            )
            return

        await self.channel_role_manager.set_role(channel_id, role)
        target_name = "スレッド" if is_thread else "チャンネル"
        await interaction.response.send_message(
            f"{interaction.user.mention} がこの{target_name}のChatbotの役割を `{role.value}` に変更しました。"
        )

    @app_commands.command(name="set_chatbot_shadow_mode", description="このチャンネルのChatbotシャドーモードを変更します")
    async def set_chatbot_shadow_mode(self, interaction: discord.Interaction, *, enabled: bool) -> None:
        """雑談の自発反応を投稿せず候補として保存する設定を変更します。"""
        channel_id = interaction.channel_id
        if channel_id is None:
            await interaction.response.send_message("チャンネル情報を取得できませんでした。", ephemeral=True)
            return
        is_thread = isinstance(interaction.channel, discord.Thread)
        if not can_change_channel_role(is_thread=is_thread, manage_channels=interaction.permissions.manage_channels):
            await interaction.response.send_message(
                "通常チャンネルのシャドーモード変更には「チャンネルの管理」権限が必要です。",
                ephemeral=True,
            )
            return
        await self.shadow_mode_manager.set_enabled(channel_id, enabled=enabled)
        state_text = "有効" if enabled else "無効"
        await interaction.response.send_message(f"このチャンネルのChatbotシャドーモードを{state_text}にしました。")

    @app_commands.command(name="export_chatbot_shadow_candidates", description="未評価のChatbotシャドー候補をExcelで出力します")
    async def export_chatbot_shadow_candidates(self, interaction: discord.Interaction) -> None:
        """実行した管理者が未評価の候補を最大100件、評価用Excel添付で返します。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("候補の出力には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        candidates = await self.shadow_candidate_store.get_unreviewed_candidates(interaction.user.id, limit=100)
        if not candidates:
            await interaction.response.send_message("未評価のシャドー候補はありません。", ephemeral=True)
            return
        fieldnames = [
            "trigger_message",
            "target_message",
            "conversation_context",
            "action",
            "content",
            "reaction_emoji",
            "reason",
            *SHADOW_EVALUATION_FIELDS,
            "comment",
            "created_at",
            "candidate_id",
            "channel_id",
            "trigger_message_id",
            "reply_to_message_id",
            "context_message_ids",
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "候補評価"
        worksheet.append([SHADOW_REVIEW_HEADERS[field_name] for field_name in fieldnames])
        worksheet.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.font = Font(name="Meiryo UI", color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for candidate in candidates:
            row = {
                "candidate_id": str(candidate.id),
                "created_at": candidate.created_at.isoformat(),
                "channel_id": candidate.channel_id,
                "trigger_message_id": candidate.trigger_message_id,
                "action": SHADOW_ACTION_LABELS.get(candidate.action, candidate.action),
                "reply_to_message_id": candidate.reply_to_message_id or "",
                "content": candidate.content,
                "reaction_emoji": candidate.reaction_emoji or "",
                "reason": SHADOW_REASON_LABELS.get(candidate.reason, candidate.reason),
                "context_message_ids": ",".join(str(message_id) for message_id in candidate.context_message_ids),
                "trigger_message": "",
                "target_message": "",
                "conversation_context": "",
                **dict.fromkeys(SHADOW_EVALUATION_FIELDS, ""),
                "comment": "",
            }
            worksheet.append([row[field] for field in fieldnames])
            context_column = fieldnames.index("conversation_context") + 1
            worksheet.cell(worksheet.max_row, context_column).value = self._format_shadow_conversation_context(
                candidate.context_snapshot
            )
            trigger_column = fieldnames.index("trigger_message") + 1
            worksheet.cell(worksheet.max_row, trigger_column).value = self._format_shadow_context_message_rich(
                candidate.context_snapshot,
                candidate.trigger_message_id,
            )
            target_column = fieldnames.index("target_message") + 1
            worksheet.cell(worksheet.max_row, target_column).value = self._format_shadow_context_message_rich(
                candidate.context_snapshot,
                candidate.reply_to_message_id,
            )
        worksheet.auto_filter.ref = worksheet.dimensions
        widths = {"A": 32, "B": 32, "C": 72, "D": 16, "E": 36, "F": 14, "G": 20, "N": 20, "O": 32}
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Meiryo UI")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        evaluation_validation = DataValidation(type="list", formula1='"◯,\u00d7,△"', allow_blank=False)
        worksheet.add_data_validation(evaluation_validation)
        for index, field_name in enumerate(fieldnames, start=1):
            if field_name in SHADOW_EVALUATION_FIELDS:
                evaluation_validation.add(
                    f"{worksheet.cell(1, index).column_letter}2:{worksheet.cell(1, index).column_letter}{len(candidates) + 1}"
                )
        worksheet.conditional_formatting.add(
            f"H2:M{len(candidates) + 1}",
            FormulaRule(formula=['H2="\u00d7"'], fill=PatternFill("solid", fgColor="F4CCCC")),
        )
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime("%Y%m%d_%H%M%S")
        await interaction.response.send_message(
            f"未評価のシャドー候補 {len(candidates)} 件を出力しました。",
            ephemeral=True,
            file=discord.File(output, filename=f"chatbot_shadow_candidates_{timestamp}.xlsx"),
        )

    @app_commands.command(
        name="import_chatbot_shadow_reviews",
        description="評価済みのChatbotシャドー候補Excelを取り込みます",
    )
    async def import_chatbot_shadow_evaluations(
        self,
        interaction: discord.Interaction,
        attachment: discord.Attachment,
    ) -> None:
        """Excelの有効な評価行だけを保存し、同じ管理者の既存評価を上書きします。"""
        if not interaction.permissions.manage_guild:
            await interaction.response.send_message("評価の取込には「サーバー管理」権限が必要です。", ephemeral=True)
            return
        workbook = load_workbook(io.BytesIO(await attachment.read()), data_only=True)
        worksheet = workbook["候補評価"]
        headers = [cell.value for cell in worksheet[1]]
        internal_headers = {display_name: field_name for field_name, display_name in SHADOW_REVIEW_HEADERS.items()}
        imported_rows = 0
        invalid_rows: list[int] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {
                internal_headers.get(str(header), str(header)): "" if value is None else str(value)
                for header, value in zip(headers, values, strict=True)
            }
            try:
                evaluation = self._parse_shadow_evaluation_row(row)
            except ValueError:
                invalid_rows.append(row_number)
                continue
            await self.shadow_candidate_store.save_evaluation(interaction.user.id, evaluation)
            imported_rows += 1
        invalid_text = "" if not invalid_rows else f" 無効な行: {', '.join(map(str, invalid_rows))}。"
        await interaction.response.send_message(f"評価を {imported_rows} 件取り込みました。{invalid_text}", ephemeral=True)

    def _parse_shadow_evaluation_row(self, row: dict[str, str]) -> ShadowEvaluationInput:
        """CSVの1行を検証し、保存用の評価データへ変換します。"""
        if any(row.get(field) not in SHADOW_EVALUATION_VALUES for field in SHADOW_EVALUATION_FIELDS):
            raise ValueError
        try:
            candidate_id = uuid.UUID(row["candidate_id"])
        except (KeyError, ValueError) as error:
            raise ValueError from error
        return ShadowEvaluationInput(
            candidate_id=candidate_id,
            **{field: row[field] for field in SHADOW_EVALUATION_FIELDS},
            issue_category="",
            comment=row.get("comment", ""),
        )

    def _format_shadow_context_message(
        self,
        context_snapshot: list[dict[str, object]],
        message_id: int | None,
    ) -> str:
        """保存済みの文脈から、CSV表示用の発言を整形します。"""
        if message_id is None:
            return ""
        for message in context_snapshot:
            if message["message_id"] == message_id:
                return f"{message['author_name']}: {message['content']}"
        return ""

    def _format_shadow_context_message_rich(
        self,
        context_snapshot: list[dict[str, object]],
        message_id: int | None,
    ) -> CellRichText:
        """反応元または反応対象を、発言者名を太字にしたExcel用リッチテキストへ変換します。"""
        if message_id is None:
            return CellRichText()
        for message in context_snapshot:
            if message["message_id"] == message_id:
                return CellRichText(
                    TextBlock(InlineFont(rFont="Meiryo UI", b=True), f"{message['author_name']}: "),
                    TextBlock(InlineFont(rFont="Meiryo UI"), str(message["content"])),
                )
        return CellRichText()

    def _format_shadow_conversation_context(self, context_snapshot: list[dict[str, object]]) -> CellRichText:
        """会話抜粋を、発言者名を太字にしたExcel用リッチテキストへ変換します。"""
        context = CellRichText()
        for index, message in enumerate(context_snapshot):
            author_font = InlineFont(rFont="Meiryo UI", b=True)
            content_font = InlineFont(rFont="Meiryo UI")
            context.append(TextBlock(author_font, f"{message['author_name']}: "))
            suffix = "\n" if index < len(context_snapshot) - 1 else ""
            context.append(TextBlock(content_font, f"{message['content']}{suffix}"))
        return context

    @app_commands.command(name="reset_chatbot_role", description="このチャンネル固有のChatbot役割を解除します")
    async def reset_chatbot_role(self, interaction: discord.Interaction) -> None:
        channel_id = interaction.channel_id
        if channel_id is None:
            await interaction.response.send_message("チャンネル情報を取得できませんでした。", ephemeral=True)
            return

        is_thread = isinstance(interaction.channel, discord.Thread)
        if not can_change_channel_role(
            is_thread=is_thread,
            manage_channels=interaction.permissions.manage_channels,
        ):
            await interaction.response.send_message(
                "通常チャンネルの役割変更には「チャンネルの管理」権限が必要です。",
                ephemeral=True,
            )
            return

        await self.channel_role_manager.clear_role(channel_id)
        role = await self.channel_role_manager.get_role(
            channel_id,
            interaction.channel.parent_id if is_thread else None,
        )
        target_name = "スレッド" if is_thread else "チャンネル"
        source = "親チャンネルから継承" if is_thread else "既定値を使用"
        await interaction.response.send_message(
            f"{interaction.user.mention} がこの{target_name}固有の設定を解除しました。"
            f"現在は `{role.value}` です。設定元: {source}。"
        )

    @commands.Cog.listener()
    async def on_message(self, message: Message) -> None:
        # 1. 明示的に呼ばれる前の会話も保持するため、チャンネルごとにパイプラインを遅延初期化
        state = await self._ensure_channel_state(message.channel.id)

        # 2. 同じチャンネルで回答を生成中の場合、生成中の文脈を変えないようメッセージを保留
        async with state.lock:
            if state.generating:
                state.pending_messages.append(message)
            else:
                await self._append_message_to_short_term_memory(message, state)

        # 3. 回答を行うかの判定
        # 3.1 ボットのメッセージについては返信しない
        if message.author.bot:
            return

        await self.memory_extraction_queue.enqueue(message.id, message.channel.id)
        await self._schedule_memory_extraction()

        bot_user = self.bot.user
        if bot_user is None:
            return

        parent_channel_id = message.channel.parent_id if isinstance(message.channel, discord.Thread) else None
        role = await self.channel_role_manager.get_role(message.channel.id, parent_channel_id)
        mentioned_bot = any(user.id == bot_user.id for user in message.mentions)
        replied_to_bot = await self._is_reply_to_bot(message)
        is_explicit_call = mentioned_bot or replied_to_bot

        await self._cancel_unanswered_question_wait(state)
        if (
            role is ChannelRole.CHAT
            and not is_explicit_call
            and is_unaddressed_question(
                content=message.clean_content,
                is_reply=message.type == discord.MessageType.reply,
                mentioned_user_ids=[user.id for user in message.mentions],
            )
        ):
            await self._schedule_unanswered_question_wait(message, state)
            return

        response_required = should_respond(
            role,
            mentioned_bot=mentioned_bot,
            replied_to_bot=replied_to_bot,
        )
        await self._update_response_schedule(
            message if response_required else None,
            state,
            role,
            is_explicit_call=is_explicit_call,
            is_unanswered_question=False,
        )

    async def _update_response_schedule(
        self,
        response_message: Message | None,
        state: ChannelProcessingState,
        role: ChannelRole,
        *,
        is_explicit_call: bool,
        is_unanswered_question: bool,
    ) -> None:
        """返信対象を更新し、最後の人間投稿から一定時間後に生成を開始します。"""
        async with state.lock:
            if state.generating:
                if response_message is not None:
                    claim_response_slot(
                        state,
                        response_message,
                        is_explicit_call=is_explicit_call,
                        is_unanswered_question=is_unanswered_question,
                    )
                return

            if response_message is not None:
                state.debounced_response_message = response_message
                state.debounced_response_is_explicit_call = is_explicit_call
                state.debounced_response_is_unanswered_question = is_unanswered_question
            if state.debounced_response_message is None:
                return

            if state.debounce_task is not None:
                state.debounce_task.cancel()

            delay_seconds = get_response_debounce_seconds(role)
            task = asyncio.create_task(self._start_response_after_delay(state, delay_seconds))
            state.debounce_task = task

        self._background_tasks.add(task)
        task.add_done_callback(self._background_tasks.discard)

    async def _schedule_memory_extraction(self) -> None:
        """未処理投稿を一定時間まとめて長期記憶として抽出します。"""
        delay_seconds = (
            0
            if await self.memory_extraction_queue.count_pending() >= MEMORY_EXTRACTION_BATCH_SIZE
            else MEMORY_EXTRACTION_WAIT_SECONDS
        )
        if self._memory_extraction_task is not None and not self._memory_extraction_task.done():
            if delay_seconds != 0:
                return
            self._memory_extraction_task.cancel()
        task = asyncio.create_task(self._extract_long_term_memories_after_wait(delay_seconds))
        self._memory_extraction_task = task
        self._background_tasks.add(task)
        task.add_done_callback(self._background_tasks.discard)

    async def _extract_long_term_memories_after_wait(self, delay_seconds: float) -> None:
        """投稿をまとめて抽出し、失敗時はキューを再試行対象へ戻します。"""
        try:
            await asyncio.sleep(delay_seconds)
        except asyncio.CancelledError:
            return
        while True:
            queue_items = await self.memory_extraction_queue.claim_pending(MEMORY_EXTRACTION_BATCH_SIZE)
            message_ids = [item.message_id for item in queue_items]
            if not message_ids:
                return
            if not await self._extract_long_term_memory_batch(message_ids):
                return
            if len(message_ids) < MEMORY_EXTRACTION_BATCH_SIZE:
                return

    async def _extract_long_term_memory_batch(self, message_ids: list[int]) -> bool:
        """確保済みの投稿群から記憶を抽出し、処理結果をキューへ反映します。"""
        try:
            stored_messages = await self.short_term_message_store.get_by_ids(message_ids)
            messages = [
                MessageInMemory(
                    message_id=message.message_id,
                    author_id=message.author_id,
                    author_name=message.author_name,
                    content=message.content,
                    reply_to_message_id=message.reply_to_message_id,
                    mentioned_user_ids=message.mentioned_user_ids,
                    timestamp=message.created_at,
                )
                for message in stored_messages
            ]
            members = list(self.bot.get_all_members())
            member_ids = {member.id for member in members}
            active_aliases = await self.member_alias_store.get_active_aliases()
            aliases_by_user_id: dict[int, list[str]] = {}
            for alias, target_user_id in active_aliases.items():
                aliases_by_user_id.setdefault(target_user_id, []).append(alias)
            extraction = await self.long_term_memory_extractor.extract(
                messages,
                [
                    {
                        "user_id": member.id,
                        "display_name": member.display_name,
                        "username": member.name,
                        "aliases": aliases_by_user_id.get(member.id, []),
                    }
                    for member in members
                ],
            )
            messages_by_id = {message.message_id: message for message in messages}
            member_names = {
                member.id: {
                    normalize_member_alias(member.display_name),
                    normalize_member_alias(member.name),
                }
                for member in members
            }
            for alias_candidate in extraction.aliases:
                if alias_candidate.target_user_id not in member_ids:
                    continue
                normalized_alias = normalize_member_alias(alias_candidate.alias)
                if not normalized_alias or normalized_alias in member_names[alias_candidate.target_user_id]:
                    continue
                evidence = [
                    messages_by_id[message_id]
                    for message_id in alias_candidate.evidence_message_ids
                    if message_id in messages_by_id
                ]
                if not evidence:
                    continue
                await self.member_alias_store.save(
                    MemberAliasInput(
                        alias=alias_candidate.alias,
                        target_user_id=alias_candidate.target_user_id,
                        evidence_message_ids=[message.message_id for message in evidence],
                        evidence_author_ids=[message.author_id for message in evidence],
                        evidence_excerpts=[message.content for message in evidence],
                    )
                )
            active_aliases = await self.member_alias_store.get_active_aliases()
            for candidate in extraction.candidates:
                evidence = [
                    messages_by_id[message_id] for message_id in candidate.evidence_message_ids if message_id in messages_by_id
                ]
                if not evidence:
                    continue
                target_user_id = candidate.target_user_id if candidate.target_user_id in member_ids else None
                if target_user_id is None and candidate.external_entity_name:
                    target_user_id = active_aliases.get(normalize_member_alias(candidate.external_entity_name))
                external_entity_name = candidate.external_entity_name if target_user_id is None else None
                target_resolution = self._normalize_memory_target_resolution(
                    target_user_id,
                    external_entity_name,
                )
                embedding_response = await AsyncOpenAI().embeddings.create(
                    model="text-embedding-3-large",
                    input=candidate.content,
                )
                await self.long_term_memory_store.save(
                    LongTermMemoryInput(
                        target_user_id=target_user_id,
                        external_entity_name=external_entity_name,
                        target_resolution=target_resolution,
                        kind=candidate.kind,
                        content=candidate.content,
                        source_type=candidate.source_type,
                        is_sensitive=candidate.is_sensitive,
                        evidence_message_ids=[message.message_id for message in evidence],
                        evidence_author_ids=[message.author_id for message in evidence],
                        evidence_excerpts=[message.content for message in evidence],
                        embedding=embedding_response.data[0].embedding,
                    )
                )
        except Exception:
            logger.exception("Failed to extract chatbot long-term memories (message_ids=%s)", message_ids)
            await self.memory_extraction_queue.restore_pending(message_ids)
            return False
        await self.memory_extraction_queue.complete(message_ids)
        return True

    def _normalize_memory_target_resolution(
        self,
        target_user_id: int | None,
        external_entity_name: str | None,
    ) -> str:
        """記憶対象の保存形式を、実際に設定された識別子と矛盾しない形へ揃えます。"""
        if target_user_id is not None:
            return "member"
        if external_entity_name:
            return "external"
        return "unresolved"

    async def _start_response_after_delay(
        self,
        state: ChannelProcessingState,
        delay_seconds: float,
    ) -> None:
        """デバウンス時間の経過後、最新の返信対象に対する生成を開始します。"""
        try:
            await asyncio.sleep(delay_seconds)
        except asyncio.CancelledError:
            return

        async with state.lock:
            if state.debounce_task is not asyncio.current_task():
                return

            message = state.debounced_response_message
            is_explicit_call = state.debounced_response_is_explicit_call
            is_unanswered_question = state.debounced_response_is_unanswered_question
            state.debounce_task = None
            state.debounced_response_message = None
            state.debounced_response_is_explicit_call = False
            state.debounced_response_is_unanswered_question = False
            if message is None or not claim_response_slot(
                state,
                message,
                is_explicit_call=is_explicit_call,
                is_unanswered_question=is_unanswered_question,
            ):
                return

        await self._process_response_queue(
            message,
            state,
            is_explicit_call=is_explicit_call,
            is_unanswered_question=is_unanswered_question,
        )

    async def _schedule_unanswered_question_wait(self, message: Message, state: ChannelProcessingState) -> None:
        """宛先のない質問への回答を、人間の反応を優先して遅延させます。"""
        async with state.lock:
            if state.debounce_task is not None and not state.debounced_response_is_explicit_call:
                state.debounce_task.cancel()
                state.debounce_task = None
                state.debounced_response_message = None
                state.debounced_response_is_unanswered_question = False

            wait_minutes = get_unanswered_question_wait_minutes(
                self.unanswered_question_minimum_wait_minutes,
                self.unanswered_question_maximum_wait_minutes,
            )
            task = asyncio.create_task(self._answer_unanswered_question_after_wait(message, state, wait_minutes))
            state.unanswered_question_task = task
            state.unanswered_question_message_id = message.id

        self._background_tasks.add(task)
        task.add_done_callback(self._background_tasks.discard)

    async def _cancel_unanswered_question_wait(self, state: ChannelProcessingState) -> None:
        """新しい人間投稿を受けたため、待機中の宛先のない質問への回答を取り消します。"""
        async with state.lock:
            if state.unanswered_question_task is not None:
                state.unanswered_question_task.cancel()
            state.unanswered_question_task = None
            state.unanswered_question_message_id = None

    async def _answer_unanswered_question_after_wait(
        self,
        message: Message,
        state: ChannelProcessingState,
        wait_minutes: int,
    ) -> None:
        """待機後も質問が有効なら、短く答えられる場合だけ回答を生成します。"""
        try:
            await asyncio.sleep(datetime.timedelta(minutes=wait_minutes).total_seconds())
        except asyncio.CancelledError:
            return

        parent_channel_id = message.channel.parent_id if isinstance(message.channel, discord.Thread) else None
        role = await self.channel_role_manager.get_role(message.channel.id, parent_channel_id)
        if role is not ChannelRole.CHAT:
            return

        async with state.lock:
            if state.unanswered_question_task is not asyncio.current_task():
                return
            state.unanswered_question_task = None
            state.unanswered_question_message_id = None
            if not claim_response_slot(
                state,
                message,
                is_explicit_call=False,
                is_unanswered_question=True,
            ):
                return

        await self._process_response_queue(
            message,
            state,
            is_explicit_call=False,
            is_unanswered_question=True,
        )

    @commands.Cog.listener()
    async def on_message_delete(self, message: Message) -> None:
        """待機中の質問が削除された場合は、遅延した回答を取り消します。"""
        await self.short_term_message_store.delete(message.id)
        await self.memory_extraction_queue.delete(message.id)
        state = self._channel_states.get(message.channel.id)
        if state is None:
            return

        async with state.lock:
            self.response_pipelines[message.channel.id].short_term_memory.remove(message.id)
            if state.unanswered_question_message_id != message.id:
                return
            if state.unanswered_question_task is not None:
                state.unanswered_question_task.cancel()
            state.unanswered_question_task = None
            state.unanswered_question_message_id = None

    @commands.Cog.listener()
    async def on_message_edit(self, before: Message, after: Message) -> None:
        """編集された投稿を短期保存と現在の短期記憶へ反映します。"""
        state = await self._ensure_channel_state(after.channel.id)
        async with state.lock:
            short_term_memory = self.response_pipelines[after.channel.id].short_term_memory
            short_term_memory.remove(before.id)
            await short_term_memory.append(after)
            stored_message = short_term_memory.get_message(after.id)
            if stored_message is None:
                return
            await self.short_term_message_store.save(
                StoredMessageInput(
                    message_id=stored_message.message_id,
                    channel_id=after.channel.id,
                    author_id=stored_message.author_id,
                    author_name=stored_message.author_name,
                    content=stored_message.content,
                    reply_to_message_id=stored_message.reply_to_message_id,
                    mentioned_user_ids=stored_message.mentioned_user_ids,
                    created_at=stored_message.timestamp,
                    is_bot=after.author.bot,
                )
            )
            if not after.author.bot:
                await self.short_term_message_store.delete_attachments(after.id)
                for attachment in after.attachments:
                    attachment_kind = self._get_attachment_kind(attachment.content_type)
                    if attachment_kind is None:
                        continue
                    await self.short_term_message_store.save_attachment(
                        StoredAttachmentInput(
                            id=attachment.id,
                            message_id=after.id,
                            url=attachment.url,
                            filename=attachment.filename,
                            content_type=attachment.content_type,
                            kind=attachment_kind,
                        )
                    )
                    self._schedule_attachment_analysis(
                        after.id,
                        attachment.id,
                        attachment.filename,
                        attachment.url,
                        attachment_kind,
                    )

            if not after.author.bot:
                await self.memory_extraction_queue.enqueue(after.id, after.channel.id)
                await self._schedule_memory_extraction()

    async def _is_reply_to_bot(self, message: Message) -> bool:
        """受信したメッセージが、このボットの発言へのDiscord返信か判定します。

        Args:
            message: on_messageで受信した判定対象のメッセージ。

        Returns:
            返信元の投稿者がこのボットの場合はTrue。
            通常投稿または他ユーザーへの返信の場合はFalse。

        """
        reference = message.reference
        bot_user = self.bot.user
        if message.type != discord.MessageType.reply or reference is None or reference.message_id is None or bot_user is None:
            return False

        referenced_author_id = get_available_referenced_author_id(reference)
        if referenced_author_id is not None:
            return referenced_author_id == bot_user.id

        short_term_memory = self.response_pipelines[message.channel.id].short_term_memory
        referenced_author_id = short_term_memory.get_author_id(reference.message_id)
        if referenced_author_id is not None:
            return referenced_author_id == bot_user.id

        try:
            referenced_message = await message.channel.fetch_message(reference.message_id)
        except discord.HTTPException:
            logger.warning(
                "Failed to fetch replied message (message_id=%s, channel_id=%s)",
                reference.message_id,
                message.channel.id,
            )
            return False

        return referenced_message.author.id == bot_user.id

    async def _process_response_queue(
        self,
        message: Message,
        state: ChannelProcessingState,
        *,
        is_explicit_call: bool,
        is_unanswered_question: bool,
    ) -> None:
        """同一チャンネルの返信要求を順番に生成し、保留メッセージを文脈へ反映します。"""
        current_message = message
        current_is_explicit_call = is_explicit_call
        current_is_unanswered_question = is_unanswered_question
        completed_normally = False
        try:
            while True:
                now = asyncio.get_running_loop().time()
                if current_is_explicit_call or can_start_spontaneous_generation(state.last_spontaneous_action_at, now):
                    await self._generate_and_send_response(
                        current_message,
                        state,
                        is_explicit_call=current_is_explicit_call,
                        is_unanswered_question=current_is_unanswered_question,
                    )
                else:
                    if await self._is_shadow_mode_for_message(
                        current_message,
                        is_explicit_call=current_is_explicit_call,
                    ):
                        await self._save_shadow_candidate(
                            current_message,
                            LLMMessage(action=ResponseAction.SILENCE, shadow_reason=ShadowReason.COOLDOWN),
                        )
                    logger.info(
                        "Skipped spontaneous chatbot generation due to cooldown (channel_id=%s)",
                        current_message.channel.id,
                    )

                async with state.lock:
                    await self._flush_pending_messages(state)
                    next_message = state.queued_response_message
                    next_is_explicit_call = state.queued_response_is_explicit_call
                    next_is_unanswered_question = state.queued_response_is_unanswered_question
                    state.queued_response_message = None
                    state.queued_response_is_explicit_call = False
                    state.queued_response_is_unanswered_question = False
                    if next_message is None:
                        state.generating = False
                        completed_normally = True
                        return
                    current_message = next_message
                    current_is_explicit_call = next_is_explicit_call
                    current_is_unanswered_question = next_is_unanswered_question
        finally:
            if not completed_normally:
                async with state.lock:
                    await self._flush_pending_messages(state)
                    state.queued_response_message = None
                    state.queued_response_is_explicit_call = False
                    state.queued_response_is_unanswered_question = False
                    state.generating = False

    async def _flush_pending_messages(self, state: ChannelProcessingState) -> None:
        """生成中に保留したメッセージを時系列順で短期記憶へ移します。"""
        for pending_message in sorted(state.pending_messages, key=lambda pending: pending.id):
            await self._append_message_to_short_term_memory(pending_message, state)
        state.pending_messages.clear()

    async def _append_message_to_short_term_memory(self, message: Message, state: ChannelProcessingState) -> None:
        """人間投稿の長時間の空白を検出し、必要に応じて短期文脈をリセットしてから保存します。"""
        short_term_memory = self.response_pipelines[message.channel.id].short_term_memory
        if not message.author.bot and should_reset_conversation(
            state.last_human_message_timestamp,
            message.created_at,
            self.conversation_reset_minutes,
        ):
            short_term_memory.reset_for_new_conversation()

        await short_term_memory.append(message)
        stored_message = short_term_memory.get_message(message.id)
        if stored_message is not None:
            await self.short_term_message_store.save(
                StoredMessageInput(
                    message_id=stored_message.message_id,
                    channel_id=message.channel.id,
                    author_id=stored_message.author_id,
                    author_name=stored_message.author_name,
                    content=stored_message.content,
                    reply_to_message_id=stored_message.reply_to_message_id,
                    mentioned_user_ids=stored_message.mentioned_user_ids,
                    created_at=stored_message.timestamp,
                    is_bot=message.author.bot,
                )
            )
            if not message.author.bot:
                for attachment in message.attachments:
                    attachment_kind = self._get_attachment_kind(attachment.content_type)
                    if attachment_kind is None:
                        continue
                    await self.short_term_message_store.save_attachment(
                        StoredAttachmentInput(
                            id=attachment.id,
                            message_id=message.id,
                            url=attachment.url,
                            filename=attachment.filename,
                            content_type=attachment.content_type,
                            kind=attachment_kind,
                        )
                    )
                    self._schedule_attachment_analysis(
                        message.id,
                        attachment.id,
                        attachment.filename,
                        attachment.url,
                        attachment_kind,
                    )
        if not message.author.bot:
            state.last_human_message_timestamp = message.created_at

    def _get_attachment_kind(self, content_type: str | None) -> str | None:
        """短期文脈の解析対象にする添付種別を返します。"""
        if content_type in {"image/jpeg", "image/png", "image/webp", "image/gif"}:
            return "image"
        if content_type == "application/pdf":
            return "pdf"
        return None

    def _schedule_attachment_analysis(
        self,
        message_id: int,
        attachment_id: int,
        filename: str,
        url: str,
        kind: str,
    ) -> None:
        """添付内容の要約を、投稿処理を待たせずに生成します。"""
        self._update_attachment_context(
            message_id,
            AttachmentInMemory(
                attachment_id=attachment_id,
                filename=filename,
                kind=kind,
                analysis_status="pending",
            ),
        )
        task = asyncio.create_task(self._analyze_attachment(message_id, attachment_id, filename, url, kind))
        self._background_tasks.add(task)
        task.add_done_callback(self._background_tasks.discard)

    async def _analyze_attachment(
        self,
        message_id: int,
        attachment_id: int,
        filename: str,
        url: str,
        kind: str,
    ) -> None:
        """画像またはPDFを解析し、短い説明と重要テキストを保存します。"""
        content_type = "input_image" if kind == "image" else "input_file"
        content_key = "image_url" if kind == "image" else "file_url"
        analysis_input = cast(
            "Any",
            [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": (
                                "添付内容をそれぞれ100文字以内で短く要約してください。"
                                "画像やPDF内で会話の理解に重要な文字情報があれば重要テキストに抜粋し、"
                                "なければ空文字にしてください。"
                            ),
                        },
                        {"type": content_type, content_key: url},
                    ],
                }
            ],
        )
        try:
            response = await AsyncOpenAI().responses.parse(
                model="gpt-5.4-mini",
                input=analysis_input,
                text_format=AttachmentAnalysis,
            )
        except Exception:
            logger.exception("Failed to analyze chatbot attachment (attachment_id=%s)", attachment_id)
            await self.short_term_message_store.save_attachment_analysis(
                attachment_id,
                summary=None,
                important_text=None,
                status="failed",
            )
            self._update_attachment_context(
                message_id,
                AttachmentInMemory(
                    attachment_id=attachment_id,
                    filename=filename,
                    kind=kind,
                    analysis_status="failed",
                ),
            )
            return
        if response.output_parsed is None:
            logger.warning("Failed to parse chatbot attachment analysis (attachment_id=%s)", attachment_id)
            await self.short_term_message_store.save_attachment_analysis(
                attachment_id,
                summary=None,
                important_text=None,
                status="failed",
            )
            self._update_attachment_context(
                message_id,
                AttachmentInMemory(
                    attachment_id=attachment_id,
                    filename=filename,
                    kind=kind,
                    analysis_status="failed",
                ),
            )
            return
        await self.short_term_message_store.save_attachment_analysis(
            attachment_id,
            summary=self._truncate_attachment_context(response.output_parsed.summary),
            important_text=self._truncate_attachment_context(response.output_parsed.important_text),
            status="completed",
        )
        self._update_attachment_context(
            message_id,
            AttachmentInMemory(
                attachment_id=attachment_id,
                filename=filename,
                kind=kind,
                analysis_status="completed",
                summary=self._truncate_attachment_context(response.output_parsed.summary),
                important_text=self._truncate_attachment_context(response.output_parsed.important_text),
            ),
        )

    def _truncate_attachment_context(self, text: str | None) -> str | None:
        """添付の解析結果を会話文脈用の上限以内に収めます。"""
        if text is None:
            return None
        return text[:ATTACHMENT_CONTEXT_MAX_CHARACTERS]

    def _update_attachment_context(self, message_id: int, attachment: AttachmentInMemory) -> None:
        """解析完了後、稼働中の短期記憶へ添付情報を反映します。"""
        for response_pipeline in self.response_pipelines.values():
            response_pipeline.short_term_memory.set_attachment_analysis(message_id, attachment)

    async def _generate_and_send_response(
        self,
        message: Message,
        state: ChannelProcessingState,
        *,
        is_explicit_call: bool,
        is_unanswered_question: bool,
    ) -> None:
        """短期記憶から回答を生成し、生成中に文脈が更新されていなければ送信します。"""
        generation_revision = state.generation_revision
        async with message.channel.typing():
            parent_channel_id = message.channel.parent_id if isinstance(message.channel, discord.Thread) else None
            role = await self.channel_role_manager.get_role(message.channel.id, parent_channel_id)
            long_term_memory_context = await self._get_long_term_memory_context(message.channel.id)
            generated_response = await self.response_pipelines[message.channel.id].generate_response(
                role,
                is_unanswered_question=is_unanswered_question,
                long_term_memory_context=long_term_memory_context,
            )

        async with state.lock:
            if not is_generation_current(state, generation_revision):
                return
            if await self._is_shadow_mode_for_message(message, is_explicit_call=is_explicit_call):
                await self._save_shadow_candidate(message, generated_response)
                return
            await self._execute_response_action(message, generated_response, state, is_explicit_call=is_explicit_call)

    async def _get_long_term_memory_context(self, channel_id: int) -> str:
        """直近会話に意味的に近い有効記憶を応答用テキストへ整形します。"""
        try:
            short_term_memory = self.response_pipelines[channel_id].short_term_memory
            search_messages = short_term_memory.memory[-MEMORY_SEARCH_CONTEXT_MESSAGE_COUNT:]
            if not search_messages:
                return ""
            search_context = json.dumps(
                [memory_message.to_dict() for memory_message in search_messages],
                ensure_ascii=False,
            )
            latest_message_context = get_latest_memory_search_query(search_messages[-1])
            search_queries = [latest_message_context]
            if len(search_messages) > 1:
                # 短い質問の意味が周辺会話に埋もれないよう、最新投稿と会話全体を別々に検索します。
                search_queries.append(search_context)
            embedding_response = await AsyncOpenAI().embeddings.create(
                model="text-embedding-3-large",
                input=search_queries,
            )
            target_user_ids = {
                user_id
                for memory_message in search_messages
                for user_id in (memory_message.author_id, *memory_message.mentioned_user_ids)
            }
            normalized_search_context = search_context.casefold()
            for member in self.bot.get_all_members():
                known_names = {member.display_name.casefold(), member.name.casefold()}
                if any(name and name in normalized_search_context for name in known_names):
                    target_user_ids.add(member.id)
            active_aliases = await self.member_alias_store.get_active_aliases()
            target_user_ids.update(find_user_ids_by_member_alias(search_context, active_aliases))
            memories_by_id: dict[uuid.UUID, ChatbotLongTermMemory] = {}
            for embedding_data in embedding_response.data:
                matched_memories = await self.long_term_memory_store.search(
                    embedding_data.embedding,
                    target_user_ids,
                    MEMORY_SEARCH_MAXIMUM_COSINE_DISTANCE,
                    20,
                )
                for memory in matched_memories:
                    memories_by_id.setdefault(memory.id, memory)
            memories = list(memories_by_id.values())[:20]
        except Exception:
            logger.exception("Failed to search chatbot long-term memories (channel_id=%s)", channel_id)
            return ""
        logger.info(
            "Selected chatbot long-term memories (channel_id=%s, target_user_ids=%s, memory_ids=%s)",
            channel_id,
            sorted(target_user_ids),
            [str(memory.id) for memory in memories],
        )
        return "\n".join(
            f"- [target={memory.target_user_id or memory.external_entity_name or 'shared'}; "
            f"kind={memory.kind}; source={memory.source_type}] {memory.content}"
            for memory in memories
        )

    async def _execute_response_action(
        self,
        message: Message,
        response: LLMMessage,
        state: ChannelProcessingState,
        *,
        is_explicit_call: bool,
    ) -> None:
        """構造化された応答行動をDiscord上で実行します。"""
        if response.action is ResponseAction.SILENCE:
            return

        now = asyncio.get_running_loop().time()
        if not is_explicit_call and not can_execute_spontaneous_action(
            response.action,
            state.last_spontaneous_action_at,
            now,
        ):
            logger.info(
                "Skipped spontaneous chatbot action due to cooldown (action=%s, channel_id=%s)",
                response.action.value,
                message.channel.id,
            )
            return

        if response.action is ResponseAction.MESSAGE:
            await message.channel.send(response.content)
            if not is_explicit_call:
                state.last_spontaneous_action_at = now
            return

        reply_to_message_id = response.reply_to_message_id
        short_term_memory = self.response_pipelines[message.channel.id].short_term_memory
        if (
            reply_to_message_id is None
            or not short_term_memory.can_target_message(reply_to_message_id)
            or not isinstance(message.channel, discord.TextChannel | discord.Thread)
        ):
            logger.warning(
                "Generated response refers to unavailable message (action=%s, message_id=%s, channel_id=%s)",
                response.action.value,
                reply_to_message_id,
                message.channel.id,
            )
            return

        target_message = message.channel.get_partial_message(reply_to_message_id)
        if response.action is ResponseAction.REACTION:
            reaction_emoji = response.reaction_emoji
            if reaction_emoji is None:
                logger.warning("Generated reaction has no emoji (message_id=%s)", reply_to_message_id)
                return
            await target_message.add_reaction(reaction_emoji)
            if not is_explicit_call:
                state.last_spontaneous_action_at = now
            return

        await target_message.reply(response.content)
        if not is_explicit_call:
            state.last_spontaneous_action_at = now

    async def _is_shadow_mode_for_message(self, message: Message, *, is_explicit_call: bool) -> bool:
        """明示依頼以外のchat役割でシャドーモードを適用するか判定します。"""
        if is_explicit_call:
            return False
        parent_channel_id = message.channel.parent_id if isinstance(message.channel, discord.Thread) else None
        role = await self.channel_role_manager.get_role(message.channel.id, parent_channel_id)
        return role is ChannelRole.CHAT and await self.shadow_mode_manager.is_enabled(message.channel.id)

    async def _save_shadow_candidate(self, message: Message, response: LLMMessage) -> None:
        """モデルが選んだ自発反応を、投稿せず評価用候補として保存します。"""
        short_term_memory = self.response_pipelines[message.channel.id].short_term_memory
        await self.shadow_candidate_store.save(
            ShadowCandidateInput(
                channel_id=message.channel.id,
                trigger_message_id=message.id,
                action=response.action.value,
                reply_to_message_id=response.reply_to_message_id,
                content=response.content,
                reaction_emoji=response.reaction_emoji,
                reason=response.shadow_reason.value,
                context_message_ids=[memory_message.message_id for memory_message in short_term_memory.memory],
                context_snapshot=[memory_message.to_dict() for memory_message in short_term_memory.memory],
            )
        )


async def setup(bot: commands.Bot, session_factory: async_sessionmaker[AsyncSession]) -> None:
    await bot.add_cog(ChatBot(bot, session_factory))
